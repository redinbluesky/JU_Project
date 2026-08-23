"""WP-12-B PDF/XLSX/통계 정합성 + SQLite 재오픈 보존 통합 테스트.

시나리오(지시서 B)를 4개의 독립 테스트로 검증한다. 각 테스트는 기능 스코프
임시 파일 DB(tmp_path/integration.db)로 격리되며, 실제 FastAPI 라우터를
의존성 오버라이드(get_db)로 호출한다.

1. 접수 생성 → DISINFECTED 전이
2. PDF 다운로드 200 / application/pdf / %PDF 헤더 / 크기>0 / 프로토타입 표시
3. 통계 API 와 목록 API 의 동일 필터 total 일치
4. 동일 필터 XLSX 다운로드 → openpyxl 재로드(필수 10 헤더, 행수=목록, DB 값 대조,
   formula data_type='f' 셀 0개)
5. DELIVERED 까지 전이 → completion_date 확인
6. 같은 SQLite 파일을 새 engine/session 으로 재오픈 →
   Request 1건 / DELIVERED / history 4건(sequence 1..4) / completion_date 보존

- 실제 임시 .db 파일 사용 (재오픈 검증 필요)
- PDF/XLSX 는 서비스 디폴트 출력 dir 함수(_default_pdf_dir/_default_xlsx_dir)를
  tmp_path 로 monkeypatch → API 경로로 생성해도 runtime/ 미오염 (앱 코드 무수정)
- 기존 테스트 파일 미수정, 앱 코드 미수정
"""

from __future__ import annotations

import re
from datetime import date, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

import app.main.services.excel_service as excel_service
import app.main.services.pdf_service as pdf_service
from app.main.api.requests import get_db
from app.main.db.base import Base
from app.main.models.models import BusinessOffice, Request, RequestNoCounter, RequestStatusHistory
from app.main.services.excel_utils import sanitize_excel_cell
from app.server import app

PROTOTYPE_NOTICE = pdf_service.PROTOTYPE_NOTICE

OFFICE_CODE = "OFFICE_A"
OFFICE_NAME = "서울사업소"
FORMULA_ADDRESS = "=SUM(A1:A2)"  # 수식 인젝션 시드값 (sanitize 대상)
NORMAL_ADDRESS = "서울시 종로구 통합로 1"

D_A = (date.today() + timedelta(days=3)).isoformat()
D_B = (date.today() + timedelta(days=5)).isoformat()
D_C = (date.today() + timedelta(days=6)).isoformat()

EXPECTED_HEADERS = (
    "접수번호",
    "사업소",
    "수거 희망일",
    "수거 장소 유형",
    "수거 주소",
    "전동침대 수량",
    "휠체어 수량",
    "기타 소형 용구 수량",
    "상태",
    "완료일",
)


# ---------------------------------------------------------------------------
# 픽스처
# ---------------------------------------------------------------------------

@pytest.fixture
def file_db_path(tmp_path: Path) -> Path:
    """실제 임시 .db 파일 (재오픈 검증 대상). 파일만 물리 생성."""
    path = tmp_path / "integration.db"
    eng = create_engine(f"sqlite:///{path}")
    conn = eng.connect()  # 파일 생성
    conn.close()
    eng.dispose()
    return path


@pytest.fixture
def session(file_db_path: Path):
    """임시 파일 DB 세션 (teardown 시 close)."""
    engine = create_engine(
        f"sqlite:///{file_db_path}",
        echo=False,
        connect_args={"check_same_thread": False},
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autocommit=False, expire_on_commit=False)
    db = SessionLocal()

    # 시딩: 사업소 1개 + 접수번호 카운터 (create_request 의 _get_next_request_no 필요)
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)
    db.add(BusinessOffice(code=OFFICE_CODE, name=OFFICE_NAME, created_at=now))
    db.commit()
    db.add(RequestNoCounter(id=1, current_value=0))
    db.commit()

    yield db
    db.close()
    engine.dispose()


def _snap_runtime(d: Path) -> set[str]:
    return {p.name for p in d.glob("*")} if d.is_dir() else set()


@pytest.fixture
def client(session, tmp_path: Path, monkeypatch):
    """실제 app 라우터 호출용 TestClient.

    - get_db 의존성을 임시 파일 DB 세션으로 오버라이드
    - pdf/excel 서비스의 디폴트 출력 dir 함수를 tmp_path 하위로 monkeypatch
    - 테스트 전후 runtime/pdf, runtime/xlsx 파일 목록 불변 검증 (오염 금지)
    """
    pdf_out = tmp_path / "pdf_out"
    xlsx_out = tmp_path / "xlsx_out"
    pdf_out.mkdir(parents=True, exist_ok=True)
    xlsx_out.mkdir(parents=True, exist_ok=True)

    def override_get_db():
        yield session

    app.dependency_overrides[get_db] = override_get_db
    monkeypatch.setattr(pdf_service, "_default_pdf_dir", lambda: pdf_out)
    monkeypatch.setattr(excel_service, "_default_xlsx_dir", lambda: xlsx_out)

    runtime_pdf = Path("runtime/pdf")
    runtime_xlsx = Path("runtime/xlsx")
    before_pdf = _snap_runtime(runtime_pdf)
    before_xlsx = _snap_runtime(runtime_xlsx)

    c = TestClient(app)
    c.pdf_out = pdf_out
    c.xlsx_out = xlsx_out
    yield c

    app.dependency_overrides.pop(get_db, None)
    assert _snap_runtime(runtime_pdf) == before_pdf, "runtime/pdf 가 테스트에 의해 오염됨"
    assert _snap_runtime(runtime_xlsx) == before_xlsx, "runtime/xlsx 가 테스트에 의해 오염됨"


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _payload(office_id: int, pickup_date: str, address: str, eb: int, wc: int, ot: int) -> dict:
    return {
        "business_office_id": office_id,
        "pickup_date": pickup_date,
        "pickup_location_type": "자택",
        "pickup_address": address,
        "electric_bed_quantity": eb,
        "wheelchair_quantity": wc,
        "other_small_quantity": ot,
    }


_CHAIN = {
    "RECEIVED": [],
    "PICKED_UP": ["PICKED_UP"],
    "DISINFECTED": ["PICKED_UP", "DISINFECTED"],
    "DELIVERED": ["PICKED_UP", "DISINFECTED", "DELIVERED"],
}


def _create_and_transition(
    client: TestClient,
    session,
    office_id: int,
    pickup_date: str,
    address: str,
    eb: int,
    wc: int,
    ot: int,
    advance_to: str = "DISINFECTED",
) -> dict:
    """POST 생성 후 advance_to 상태까지 순차 전이. 반환: {id, request_no, status}."""
    r = client.post("/api/requests", json=_payload(office_id, pickup_date, address, eb, wc, ot))
    assert r.status_code == 201, r.text
    body = r.json()
    rid = body["id"]
    assert body["current_status"] == "RECEIVED"

    for target in _CHAIN[advance_to]:
        r = client.patch(f"/api/requests/{rid}/status", json={"target_status": target})
        assert r.status_code == 200, r.text

    req = session.get(Request, rid)
    assert req is not None
    return {"id": rid, "request_no": req.request_no, "status": req.current_status.value}


def _unescapew_pdf_bytes(data: bytes) -> bytes:
    """PDF 문자열 리터럴 이스케이프를 원본 바이트로 복원 (WP-09 검증 방식 재사용)."""
    out = bytearray()
    i = 0
    n = len(data)
    while i < n:
        ch = data[i]
        if ch == 0x5C and i + 1 < n:  # backslash
            nxt = data[i + 1]
            if 0x30 <= nxt <= 0x37:  # octal escape
                m = re.match(rb"\\([0-7]{1,3})", data[i:i + 4])
                if m:
                    out.append(int(m.group(1), 8))
                    i += m.end()
                    continue
            elif nxt in (0x5C, 0x28, 0x29):
                out.append(nxt)
                i += 2
                continue
        out.append(ch)
        i += 1
    return bytes(out)


# ---------------------------------------------------------------------------
# 테스트
# ---------------------------------------------------------------------------

def test_pdf_integration(client, session):
    """1-2. 접수 생성 → DISINFECTED 전이 → PDF 다운로드 검증."""
    a = _create_and_transition(client, session, 1, D_A, NORMAL_ADDRESS, 2, 1, 0, "DISINFECTED")

    r = client.get(f"/api/requests/{a['id']}/pdf")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/pdf"
    data = r.content
    assert data.startswith(b"%PDF")
    assert len(data) > 0

    # 1) PDF 가 tmp_path(pdf_out)에만 생성, 파일명 {request_no}.pdf, 크기>0
    files = list(client.pdf_out.glob("*.pdf"))
    assert len(files) == 1
    f = files[0]
    assert f.name == f"{a['request_no']}.pdf"
    assert f.stat().st_size > 0
    assert f.read_bytes().startswith(b"%PDF")

    # 2) 프로토타입 표시: /Title 메타에 UTF-16BE 로 기록 (WP-09 검증 방식)
    restored = _unescapew_pdf_bytes(data)
    assert PROTOTYPE_NOTICE.encode("utf-16-be") in restored, "PDF Title 에 프로토타입 표시가 없다"
    # 문서 콘텐츠에 접수번호 존재 (ASCII)
    assert a["request_no"].encode("ascii") in data

    # 3) 상태 확인
    assert a["status"] == "DISINFECTED"
    assert session.get(Request, a["id"]).current_status.value == "DISINFECTED"


def test_statistics_list_consistency(client, session):
    """3. 동일 필터 기준 통계 API 와 목록 API 의 total 일치 (전체 + status 필터)."""
    a = _create_and_transition(client, session, 1, D_A, FORMULA_ADDRESS, 2, 1, 0, "DISINFECTED")
    _create_and_transition(client, session, 1, D_B, NORMAL_ADDRESS, 0, 2, 1, "PICKED_UP")

    # 전체 (필터 없음)
    r_list = client.get("/api/requests")
    assert r_list.status_code == 200, r_list.text
    r_stat = client.get("/api/statistics")
    assert r_stat.status_code == 200, r_stat.text

    lb = r_list.json()
    sb = r_stat.json()
    assert lb["total"] == 2
    assert len(lb["items"]) == 2
    assert sb["total_requests"] == lb["total"]
    assert sb["by_status"] == {
        "RECEIVED": 0,
        "PICKED_UP": 1,
        "DISINFECTED": 1,
        "DELIVERED": 0,
    }
    items_sum = {
        "eb": sum(i["electric_bed_quantity"] for i in lb["items"]),
        "wc": sum(i["wheelchair_quantity"] for i in lb["items"]),
        "ot": sum(i["other_small_quantity"] for i in lb["items"]),
    }
    q = sb["quantities"]
    assert q["electric_bed"] == items_sum["eb"] == 2
    assert q["wheelchair"] == items_sum["wc"] == 3
    assert q["other_small"] == items_sum["ot"] == 1
    assert q["total"] == items_sum["eb"] + items_sum["wc"] + items_sum["ot"] == 6

    # 동일 필터: current_status=DISINFECTED
    r_list_f = client.get("/api/requests", params={"current_status": "DISINFECTED"})
    r_stat_f = client.get("/api/statistics", params={"current_status": "DISINFECTED"})
    assert r_list_f.status_code == 200
    assert r_stat_f.status_code == 200
    lb_f = r_list_f.json()
    sb_f = r_stat_f.json()
    assert lb_f["total"] == 1
    assert sb_f["total_requests"] == lb_f["total"]
    assert sb_f["by_status"]["DISINFECTED"] == 1
    assert len(lb_f["items"]) == 1
    assert lb_f["items"][0]["id"] == a["id"]
    assert lb_f["items"][0]["request_no"] == a["request_no"]


def test_xlsx_db_consistency(client, session):
    """4. 동일 필터 XLSX 다운로드 → openpyxl 재로드: 10헤더/행수=목록/DB 대조/수식 0."""
    a = _create_and_transition(client, session, 1, D_A, FORMULA_ADDRESS, 2, 1, 0, "DISINFECTED")
    b = _create_and_transition(client, session, 1, D_B, NORMAL_ADDRESS, 0, 2, 1, "PICKED_UP")

    r_list = client.get("/api/requests")
    assert r_list.status_code == 200
    expected_total = r_list.json()["total"]

    r_exp = client.get("/api/requests/export")
    assert r_exp.status_code == 200, r_exp.text
    assert r_exp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r_exp.content

    # tmp_path(xlsx_out)에만 1개 파일
    files = list(client.xlsx_out.glob("*.xlsx"))
    assert len(files) == 1
    wb = load_workbook(files[0], data_only=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 1) 필수 10 헤더 (첫 행, 정확한 순서)
    assert rows[0] == EXPECTED_HEADERS
    assert len(rows) == 1 + expected_total

    # 2) 행 수 = 목록 결과
    assert len(rows) - 1 == expected_total == 2

    # 3) 주요 DB 원본 값 대조 (pickup_date ASC, id ASC 정렬)
    db_rows = session.query(Request).order_by(Request.pickup_date, Request.id).all()
    assert len(db_rows) == 2
    req_a, req_b = db_rows
    assert req_a.id == a["id"]
    assert req_b.id == b["id"]

    row1 = rows[1]
    row2 = rows[2]
    # 완료일 미설정 → 빈 셀 → openpyxl values_only 는 None 반환
    assert row1 == (
        req_a.request_no,
        OFFICE_NAME,
        req_a.pickup_date.isoformat(),
        "자택",
        sanitize_excel_cell(FORMULA_ADDRESS),  # "'" 접두 → 수식 비활성화
        req_a.electric_bed_quantity,
        req_a.wheelchair_quantity,
        req_a.other_small_quantity,
        req_a.current_status.value,
        None,
    )
    assert row2 == (
        req_b.request_no,
        OFFICE_NAME,
        req_b.pickup_date.isoformat(),
        "자택",
        NORMAL_ADDRESS,
        req_b.electric_bed_quantity,
        req_b.wheelchair_quantity,
        req_b.other_small_quantity,
        req_b.current_status.value,
        None,
    )

    # 4) formula data_type='f' 셀 0개 (수식 인젝션 방지)
    all_cells = [c for row in ws.iter_rows() for c in row]
    formula_cells = [c for c in all_cells if c.data_type == "f"]
    assert formula_cells == [], (
        f"formula 셀 발견: {[(c.coordinate, c.value) for c in formula_cells]}"
    )

    # 수식 시작 문자열 주소가 텍스트로만 존재 (sanitize 가 ' 접두 부여)
    addr_cell = ws.cell(row=2, column=5)
    assert addr_cell.data_type in ("s",)
    assert str(addr_cell.value).startswith("'=SUM")


def test_sqlite_reopen_preserves_state(client, session, file_db_path: Path):
    """5-6. DELIVERED 전이 후 같은 .db 파일을 새 engine/session 으로 재오픈."""
    info = _create_and_transition(client, session, 1, D_C, NORMAL_ADDRESS, 1, 0, 1, "DELIVERED")
    rid = info["id"]
    req_before = session.get(Request, rid)
    assert req_before is not None
    assert req_before.current_status.value == "DELIVERED"
    assert req_before.completion_date is not None
    completion_before = req_before.completion_date

    # 기존 엔진/세션 종료 (재시작 시뮬레이션)
    session.close()

    # 새 engine/session 으로 재오픈 (실제 파일 기반)
    engine2 = create_engine(
        f"sqlite:///{file_db_path}",
        echo=False,
        connect_args={"check_same_thread": False},
    )
    try:
        Session2 = sessionmaker(bind=engine2, autocommit=False, expire_on_commit=False)
        s2 = Session2()
        reqs = s2.query(Request).all()
        assert len(reqs) == 1, f"재오픈 후 Request 건수 불일치: {len(reqs)}"
        req = reqs[0]
        assert req.id == rid
        assert req.request_no == info["request_no"]
        assert req.current_status.value == "DELIVERED"
        assert req.completion_date is not None
        assert req.completion_date == completion_before

        hist = (
            s2.query(RequestStatusHistory)
            .filter(RequestStatusHistory.request_id == rid)
            .order_by(RequestStatusHistory.sequence)
            .all()
        )
        assert len(hist) == 4, f"이력 건수 불일치: {len(hist)}"
        assert [h.sequence for h in hist] == [1, 2, 3, 4]
        assert [h.status.value for h in hist] == [
            "RECEIVED",
            "PICKED_UP",
            "DISINFECTED",
            "DELIVERED",
        ]
        s2.close()
    finally:
        engine2.dispose()

    # SQLite sidecar 파일 정리 (WAL/SHM)
    for suffix in ("-wal", "-shm"):
        side = file_db_path.with_name(file_db_path.name + suffix)
        if side.exists():
            side.unlink()
