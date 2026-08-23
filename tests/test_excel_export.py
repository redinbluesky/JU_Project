"""
엑셀 수식 인젝션 방지 검증 테스트.
sanitize_excel_cell 함수가 수식처럼 보이는 입력을 안전하게 문자열로 변환하는지 확인.
"""

import pytest
from openpyxl import Workbook, load_workbook
import tempfile
import os

from app.main.services.excel_utils import sanitize_excel_cell


class TestSanitizeExcelCell:
    """sanitize_excel_cell 함수 단위 테스트"""

    def test_equals_prefix(self):
        """=로 시작하는 값은 앞에 '가 붙어야 함"""
        result = sanitize_excel_cell('=SUM(1+1)')
        assert result == "'=SUM(1+1)"
        assert result[0] == "'"

    def test_plus_prefix(self):
        """+로 시작하는 값은 앞에 '가 붙어야 함"""
        result = sanitize_excel_cell('+1+1')
        assert result == "'+1+1"

    def test_minus_prefix(self):
        """-로 시작하는 값은 앞에 '가 붙어야 함"""
        result = sanitize_excel_cell('-1+1')
        assert result == "'-1+1"

    def test_at_prefix(self):
        """@로 시작하는 값은 앞에 '가 붙어야 함"""
        result = sanitize_excel_cell('@cmd')
        assert result == "'@cmd"

    def test_normal_string(self):
        """일반 문자열은 변경 없음"""
        result = sanitize_excel_cell('hello')
        assert result == 'hello'

    def test_empty_string(self):
        """빈 문자열은 변경 없음"""
        result = sanitize_excel_cell('')
        assert result == ''

    def test_none_value(self):
        """None은 변경 없음"""
        result = sanitize_excel_cell(None)
        assert result is None

    def test_numeric_value(self):
        """숫자는 변경 없음"""
        result = sanitize_excel_cell(123)
        assert result == 123


class TestOpenpyxlFormulaPrevention:
    """openpyxl로 실제 저장 후 data_type 검증"""

    def test_formula_without_sanitization(self):
        """sanitization 없이 =로 시작하는 값은 수식으로 저장됨 (data_type='f')"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=SUM(1+1)'
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            # data_type이 'f'(formula)여야 함 (이것이 문제)
            assert ws2['A1'].data_type == 'f', f"예상 'f', 실제 '{ws2['A1'].data_type}'"
        finally:
            os.unlink(path)

    def test_formula_with_sanitization(self):
        """sanitization 후 =로 시작하는 값은 문자열로 저장됨 (data_type='s')"""
        wb = Workbook()
        ws = wb.active
        sanitized = sanitize_excel_cell('=SUM(1+1)')
        ws['A1'] = sanitized
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            # data_type이 's'(string)여야 함 (안전)
            assert ws2['A1'].data_type == 's', f"예상 's', 실제 '{ws2['A1'].data_type}'"
            # 값도 확인 (small quote 제거 후 원래 값)
            assert ws2['A1'].value == "'=SUM(1+1)"
        finally:
            os.unlink(path)

    def test_plus_with_sanitization(self):
        """+로 시작하는 값도 sanitization 후 문자열로 저장"""
        wb = Workbook()
        ws = wb.active
        sanitized = sanitize_excel_cell('+1+1')
        ws['A1'] = sanitized
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            assert ws2['A1'].data_type == 's', f"예상 's', 실제 '{ws2['A1'].data_type}'"
        finally:
            os.unlink(path)

    def test_minus_with_sanitization(self):
        """-로 시작하는 값도 sanitization 후 문자열로 저장"""
        wb = Workbook()
        ws = wb.active
        sanitized = sanitize_excel_cell('-1+1')
        ws['A1'] = sanitized
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            assert ws2['A1'].data_type == 's', f"예상 's', 실제 '{ws2['A1'].data_type}'"
        finally:
            os.unlink(path)

    def test_at_with_sanitization(self):
        """@로 시작하는 값도 sanitization 후 문자열로 저장"""
        wb = Workbook()
        ws = wb.active
        sanitized = sanitize_excel_cell('@cmd')
        ws['A1'] = sanitized
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            assert ws2['A1'].data_type == 's', f"예상 's', 실제 '{ws2['A1'].data_type}'"
        finally:
            os.unlink(path)

    def test_full_excel_export_with_sanitization(self):
        """전체 엑셀 내보내기 시나리오: 헤더 + 데이터 + 수식인젝션테스트"""
        wb = Workbook()
        ws = wb.active
        
        # 헤더
        headers = ['접수번호', '사업소', '수거희망일', '수거장소유형', '수거주소',
                   '전동침대수량', '휠체어수량', '기타소형수량', '상태', '완료일']
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        
        # 일반 데이터
        ws.append(['10001', '서울사업소', '2026-08-25', '주택', '서울시 강남구', 1, 0, 2, '완료', '2026-08-24'])
        ws.append(['10002', '부산사업소', '2026-08-26', '아파트', '부산시 해운대구', 0, 1, 1, '대기', None])
        
        # 수식 인젝션 테스트 (sanitized)
        ws.append([
            sanitize_excel_cell('=SUM(1+1)'),
            sanitize_excel_cell('XSS 테스트'),
            sanitize_excel_cell('=EXEC()'),
            sanitize_excel_cell('danger'),
            sanitize_excel_cell('test')
        ])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            ws2 = wb2.active
            
            # A4 셀 (=SUM(1+1))이 문자열인지 확인
            assert ws2['A4'].data_type == 's', f"예상 's', 실제 '{ws2['A4'].data_type}'"
            # C4 셀 (=EXEC())이 문자열인지 확인
            assert ws2['C4'].data_type == 's', f"예상 's', 실제 '{ws2['C4'].data_type}'"
        finally:
            os.unlink(path)
