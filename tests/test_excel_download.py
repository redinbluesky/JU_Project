"""WP-11 엑셀 다운로드 API 테스트.

- 임시 SQLite(in-memory) + StaticPool 단일 연결 + TestClient + get_db 오버라이드
- 사업소 2개 + 서로 다른 pickup_date/상태를 가진 접수 4건 시딩
  (수거 주소가 수식 인젝션 페이로드 =, +, -, @ 로 시작하는 4가지 케이스)
- API 경유 XLSX 생성은 tmp_path 로 리다이렉트 (실제 runtime/xlsx/ 오염 금지)
- 응답 바이트를 openpyxl 로 직접 재로드해 헤더/행수/셀값/data_type 검증

시딩 설계 (생성 순서 id 1..4, pickup_date 순서 = id 순서):
  r1: office 1, D1, RECEIVED,  address "=SUM(1+1) 시딩주소1"
  r2: office 1, D2, PICKED_UP, address "+cmd 시딩주소2"
  r3: office 2, D3, DISINFECTED, address "-cmd 시딩주소3"
  r4: office 2, D4, DELIVERED, address "@cmd 시딩주소4" (completion_date 있음)
"""

import io
from datetime import date, datetime, timedelta, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, pool, select
from sqlalchemy.orm import sessionmaker

from app.main.db.base import Base
from app.main.models.models import BusinessOffice, Request, RequestNoCounter, RequestStatus
from app.main.api import requests as api_module
from app.main.api.requests import get_db
from app.main.services.excel_utils import sanitize_excel_cell
from app.server import app

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EXPECTED_HEADERS = [
    "접수번호",
    "사업소",
    "수거 희망일",
    "수거 장소 유형",
    "수거 주소",
    "전동침대 수량",
    "휠체어 수량",
    "기타 소형 용구 수량",
    "상태",
    "완료일",
]

OFFICE_NAMES = {1: "서울사업소", 2: "부산사업소"}

# 서로 다른 수거희망일 (오늘 기준 미래) — id 순서와 동일
D1 = (date.today() + timedelta(days=7)).isoformat()
D2 = (date.today() + timedelta(days=10)).isoformat()
D3 = (date.today() + timedelta(days=12)).isoformat()
D4 = (date.today() + timedelta(days=20)).isoformat()

# 수식 인젝션 페이로드로 시작하는 주소 (=, +, -, @)
INJECTION_ADDRESSES = {
    "r1": "=SUM(1+1) 시딩주소1",
    "r2": "+cmd 시딩주소2",
    "r3": "-cmd 시딩주소3",
    "r4": "@cmd 시딩주소4",
}


@pytest.fixture
def test_db():
    """임시 SQLite 세션 + 사업소 2개/카운터 시딩 (StaticPool 단일 연결 공유)."""
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=pool.StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autocommit=False, expire_on_commit=False)
    db = SessionLocal()

    now = datetime.now(timezone.utc)
    db.add(BusinessOffice(code="OFFICE_A", name="서울사업소", created_at=now))
    db.add(BusinessOffice(code="OFFICE_B", name="부산사업소", created_at=now))
    db.commit()
    db.add(RequestNoCounter(id=1, current_value=0))
    db.commit()

    yield db
    db.close()


@pytest.fixture
def client(test_db):
    """get_db를 임시 세션으로 오버라이드한 TestClient."""
    def override_get_db():
        yield test_db

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.pop(get_db, None)


@pytest.fixture(autouse=True)
def xlsx_dir(tmp_path, monkeypatch):
    """모든 테스트의 API 경유 XLSX 생성을 tmp_path 로 리다이렉트 (autouse).

    기본 출력 위치(runtime/xlsx/)는 개발 실행 전용으로 유지하고,
    테스트가 절대 오염시키지 못하게 한다.
    """
    real = api_module.generate_requests_xlsx

    def fake(db, filters, out_dir=None):
        return real(db, filters, out_dir=tmp_path)

    monkeypatch.setattr(api_module, "generate_requests_xlsx", fake)
    return tmp_path


def _payload(office_id: int, pickup_date: str, address: str, eb: int, wc: int, os_: int) -> dict:
    return {
        "business_office_id": office_id,
        "pickup_date": pickup_date,
        "pickup_location_type": "자택",
        "pickup_address": address,
        "electric_bed_quantity": eb,
        "wheelchair_quantity": wc,
        "other_small_quantity": os_,
    }


@pytest.fixture
def seeded(client):
    """접수 4건 시딩 + 상태 전이로 4개 상태 확보.

    r1: office 1, D1, RECEIVED
    r2: office 1, D2, PICKED_UP
    r3: office 2, D3, DISINFECTED
    r4: office 2, D4, DELIVERED
    """
    ids = {}
    specs = (
        ("r1", 1, D1, 1, 2, 3),
        ("r2", 1, D2, 4, 5, 0),
        ("r3", 2, D3, 7, 0, 2),
        ("r4", 2, D4, 0, 1, 5),
    )
    for i, (key, office, d, eb, wc, os_) in enumerate(specs, start=1):
        r = client.post(
            "/api/requests",
            json=_payload(office, d, INJECTION_ADDRESSES[key], eb, wc, os_),
        )
        assert r.status_code == 201, r.text
        ids[key] = r.json()["id"]

    r = client.patch(f"/api/requests/{ids['r2']}/status", json={"target_status": "PICKED_UP"})
    assert r.status_code == 200, r.text
    r = client.patch(f"/api/requests/{ids['r3']}/status", json={"target_status": "PICKED_UP"})
    assert r.status_code == 200, r.text
    r = client.patch(f"/api/requests/{ids['r3']}/status", json={"target_status": "DISINFECTED"})
    assert r.status_code == 200, r.text
    r = client.patch(f"/api/requests/{ids['r4']}/status", json={"target_status": "PICKED_UP"})
    assert r.status_code == 200, r.text
    r = client.patch(f"/api/requests/{ids['r4']}/status", json={"target_status": "DISINFECTED"})
    assert r.status_code == 200, r.text
    r = client.patch(f"/api/requests/{ids['r4']}/status", json={"target_status": "DELIVERED"})
    assert r.status_code == 200, r.text

    return ids


def _export(client, **params):
    r = client.get("/api/requests/export", params=params)
    assert r.status_code == 200, r.text
    return r


def _sheet(response) -> "ws":
    """응답 바이트를 openpyxl 로 재로드해 활성 시트 반환."""
    wb = load_workbook(io.BytesIO(response.content))
    return wb.active


def _data_rows(ws):
    """헤더 제외 데이터 행 (tuple 리스트)."""
    return [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]


def _expected_order(db) -> list[Request]:
    """목록 API 정렬(pickup_date ASC, id ASC)대로 DB 행 직접 조회."""
    return list(db.execute(
        select(Request).order_by(Request.pickup_date.asc(), Request.id.asc())
    ).scalars().all())


def _enum_value(v):
    return v.value if isinstance(v, (RequestStatus,)) and hasattr(v, "value") else v


class TestBasicDownload:
    """1. 기본 다운로드 200 + XLSX content-type"""

    def test_export_200_and_content_type(self, client, seeded, xlsx_dir):
        r = _export(client)
        assert r.status_code == 200
        assert r.headers["content-type"] == XLSX_CONTENT_TYPE
        # 유효한 XLSX(zip) 마직
        assert r.content.startswith(b"PK")
        assert len(r.content) > 0
        # API 생성 파일은 주입된 tmp 디렉터리에만 존재
        files = list(xlsx_dir.glob("*.xlsx"))
        assert len(files) == 1
        assert files[0].stat().st_size > 0

    def test_export_does_not_pollute_default_runtime_dir(self, client, seeded, xlsx_dir):
        """기본 출력 위치(runtime/xlsx/)가 테스트로 오염되지 않음."""
        runtime_xlsx = xlsx_dir  # 실제 경로는 프로젝트 루트 runtime/xlsx
        from pathlib import Path
        default_dir = Path(__file__).resolve().parents[1] / "runtime" / "xlsx"
        before = list(default_dir.glob("*.xlsx")) if default_dir.exists() else []
        r = _export(client)
        assert r.status_code == 200
        after = list(default_dir.glob("*.xlsx")) if default_dir.exists() else []
        assert after == before
        # 파일은 tmp 쪽에만 생성됨
        assert len(list(xlsx_dir.glob("*.xlsx"))) == 1


class TestHeaderColumns:
    """2. 헤더 10개 이름·순서 정확성"""

    def test_header_names_and_order(self, client, seeded, xlsx_dir):
        r = _export(client)
        ws = _sheet(r)
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header == tuple(EXPECTED_HEADERS)
        assert len(header) == 10
        assert ws.max_column >= 10


class TestRowCountMatchesList:
    """3. 행 수 = 목록 필터 결과 행 수"""

    def test_row_count_various_filters(self, client, seeded):
        cases = [
            ({}, 4),
            ({"business_office_id": 1}, 2),
            ({"business_office_id": 2}, 2),
            ({"current_status": "RECEIVED"}, 1),
            ({"pickup_date_from": D1, "pickup_date_to": D3}, 3),
            ({"business_office_id": 2, "current_status": "DISINFECTED"}, 1),
        ]
        for params, expected in cases:
            listing = client.get("/api/requests", params=params)
            assert listing.status_code == 200, listing.text
            r = _export(client, **params)
            ws = _sheet(r)
            rows = _data_rows(ws)
            assert len(rows) == expected
            assert len(rows) == listing.json()["total"]


class TestDateRangeOnPickupDate:
    """4. 기간 필터는 pickup_date 기준"""

    def test_date_range_filters_by_pickup_date(self, client, seeded):
        # D1~D3: r1(D1), r2(D2), r3(D3)만, r4(D4) 제외
        r = _export(client, pickup_date_from=D1, pickup_date_to=D3)
        ws = _sheet(r)
        rows = _data_rows(ws)
        dates = [row[2] for row in rows]
        assert dates == [D1, D2, D3]
        request_nos = {row[0] for row in rows}
        listing = client.get(
            "/api/requests", params={"pickup_date_from": D1, "pickup_date_to": D3}
        )
        assert {i["request_no"] for i in listing.json()["items"]} == request_nos

    def test_single_day_range(self, client, seeded):
        r = _export(client, pickup_date_from=D2, pickup_date_to=D2)
        rows = _data_rows(_sheet(r))
        assert len(rows) == 1
        assert rows[0][2] == D2  # 수거 희망일 열이 D2

    def test_from_after_to_422(self, client, seeded):
        r = client.get(
            "/api/requests/export",
            params={"pickup_date_from": D2, "pickup_date_to": D1},
        )
        assert r.status_code == 422, r.text


class TestFilterConsistency:
    """5. 사업소·상태·복합 필터 정합성 (목록 API와 동일 행 집합)"""

    def test_office_filter(self, client, seeded):
        r = _export(client, business_office_id=2)
        rows = _data_rows(_sheet(r))
        listing = client.get("/api/requests", params={"business_office_id": 2})
        assert len(rows) == listing.json()["total"] == 2
        # office 2 = r3, r4 → 서울사업소/부산사업소 중 부산사업소만
        assert {row[1] for row in rows} == {"부산사업소"}
        assert {row[0] for row in rows} == {
            i["request_no"] for i in listing.json()["items"]
        }

    def test_status_filter(self, client, seeded):
        r = _export(client, current_status="DISINFECTED")
        rows = _data_rows(_sheet(r))
        listing = client.get("/api/requests", params={"current_status": "DISINFECTED"})
        assert len(rows) == listing.json()["total"] == 1
        assert rows[0][8] == "DISINFECTED"  # 상태 열

    def test_combined_filter(self, client, seeded):
        # office 2 + DISINFECTED → r3 1건만
        params = {"business_office_id": 2, "current_status": "DISINFECTED"}
        r = _export(client, **params)
        rows = _data_rows(_sheet(r))
        listing = client.get("/api/requests", params=params)
        assert len(rows) == listing.json()["total"] == 1
        assert rows[0][1] == "부산사업소"
        assert rows[0][8] == "DISINFECTED"
        assert rows[0][0] == listing.json()["items"][0]["request_no"]


class TestZeroResult:
    """6. 0건 결과도 헤더 10개인 유효한 XLSX"""

    def test_zero_rows_still_valid_xlsx_with_header(self, client, seeded):
        far_from = (date.today() + timedelta(days=365)).isoformat()
        far_to = (date.today() + timedelta(days=366)).isoformat()
        r = _export(client, pickup_date_from=far_from, pickup_date_to=far_to)
        assert r.status_code == 200
        assert r.headers["content-type"] == XLSX_CONTENT_TYPE
        ws = _sheet(r)
        assert ws.max_row == 1  # 헤더 1행만
        assert ws.max_column >= 10
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header == tuple(EXPECTED_HEADERS)
        assert _data_rows(ws) == []


class TestValuesMatchDb:
    """7. DB 원본 값과 XLSX 각 열 값 대조"""

    def test_all_cells_match_db(self, client, seeded, test_db):
        r = _export(client)
        ws = _sheet(r)
        rows = _data_rows(ws)
        expected = _expected_order(test_db)
        assert len(rows) == len(expected) == 4

        for row, req in zip(rows, expected):
            assert row[0] == req.request_no
            assert row[1] == OFFICE_NAMES[req.business_office_id]  # id가 아니라 name
            assert row[2] == req.pickup_date.isoformat()
            loc = req.pickup_location_type.value if hasattr(req.pickup_location_type, "value") else req.pickup_location_type
            assert row[3] == loc
            # 사용자 입력 문자열은 sanitize_excel_cell() 적용 후 저장
            assert row[4] == sanitize_excel_cell(req.pickup_address)
            assert row[5] == req.electric_bed_quantity
            assert row[6] == req.wheelchair_quantity
            assert row[7] == req.other_small_quantity
            st = req.current_status.value if hasattr(req.current_status, "value") else req.current_status
            assert row[8] == st
            # 완료일 미설정 시 빈 셀(재로드 시 None)
            expected_completion = (
                req.completion_date.isoformat() if req.completion_date is not None else None
            )
            assert row[9] == expected_completion

    def test_delivered_has_completion_date_others_empty(self, client, seeded, test_db):
        r = _export(client)
        rows = _data_rows(_sheet(r))
        expected = _expected_order(test_db)
        for row, req in zip(rows, expected):
            if req.current_status == RequestStatus.DELIVERED:
                assert row[9] is not None
            else:
                assert row[9] is None


class TestFormulaInjection:
    """8. =, +, -, @ 시작 주소 저장 후 재로드 시 data_type='s' (수식 아님)"""

    def test_injection_addresses_stored_as_string(self, client, seeded, test_db):
        r = _export(client)
        ws = _sheet(r)
        expected = _expected_order(test_db)
        assert len(expected) == 4
        for i, req in enumerate(expected):
            cell = ws.cell(row=i + 2, column=5)  # 수거 주소 열
            assert cell.data_type == "s", f"{req.request_no} 주소가 문자열이 아님: {cell.data_type}"
            assert cell.value == sanitize_excel_cell(req.pickup_address)

    def test_no_formula_cells_anywhere(self, client, seeded):
        """전체 시트에 data_type='f'(수식) 셀이 없어야 한다."""
        r = _export(client)
        ws = _sheet(r)
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", f"수식 셀 발견: {cell.coordinate}"
        # 특히 =SUM(1+1) 페이로드가 수식으로 해석되지 않음
        addrs = [ws.cell(row=i, column=5).value for i in range(2, 6)]
        assert all(isinstance(a, str) and a.startswith("'") for a in addrs)


class TestExistingApisRegression:
    """9. 기존 API 회귀 (export 라우터가 int 경로/기존 라우터를 가리지 않음)"""

    def test_list_api_still_works(self, client, seeded):
        r = client.get("/api/requests")
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["total"] == 4
        assert len(body["items"]) == 4

    def test_detail_api_not_shadowed_by_export(self, client, seeded):
        # /api/requests/{int} 가 /api/requests/export 에 가로막히지 않음
        r = client.get(f"/api/requests/{seeded['r1']}")
        assert r.status_code == 200, r.text
        assert r.json()["current_status"] == "RECEIVED"
        r = client.get("/api/requests/99999")
        assert r.status_code == 404, r.text

    def test_create_still_works(self, client, seeded):
        r = client.post("/api/requests", json=_payload(1, D4, "회귀주소", 1, 0, 1))
        assert r.status_code == 201, r.text
        assert r.json()["current_status"] == "RECEIVED"

    def test_statistics_still_works(self, client, seeded):
        r = client.get("/api/statistics")
        assert r.status_code == 200, r.text
        assert r.json()["total_requests"] == 4
